from operator import truediv
from dotenv import load_dotenv, find_dotenv
from hubspot3 import Hubspot3
import os
import hubspot
from pprint import pformat, pprint
from hubspot.crm.companies import ApiException
from hubspot.crm.products import SimplePublicObjectInput, ApiException
import requests
from tkinter import *
from tkinter import filedialog
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from hubspot import HubSpot
import itertools
import time

load_dotenv() 
# t = time.localtime()
client = HubSpot(access_token=os.getenv("ACCESS_TOKEN"))

def get_products():
    try:
        api_response = client.crm.products.basic_api.get_page(limit=10, archived=False)
        pprint(api_response)
    except ApiException as e:
        print("Exception when calling basic_api->get_page: %s\n" % e)

def print_data():
    directory = filedialog.askopenfilename(initialdir="C:/", title="select company file")
    company_dict = data_to_dict("company", directory)    
    for key, value in (company_dict).items():
        print(key, ' - ', value)
   
def get_name(id):
    url = "https://api.hubapi.com/companies/v2/companies/"+ str(id)
    headers = {
        'Content-Type': "application/json"
        }
    response = requests.get(url=url)
    response = response.text # makes response searchable
    x = response.find('"name":{"value":"') #looks for name in response and finds value it starts
    snippet = response[x:(x+150)] # encompasses area of response where name should be located
    company_name = snippet[17:snippet.find('","')] # filters thru snippet to get company name
    return company_name

def load_to_text(companies):
    with open("data.txt", "r+") as data:
        for company,id in companies.items():
            line = company + '-' + str(id) + '\n'
            print(line)
            data.write(line)  
    print("Data loaded succesfully")
    data.close

def makeParent_companyToCompany(parent_id, child_id):
    try:
        api_response = client.crm.companies.associations_api.create(
            company_id = parent_id, 
            to_object_type = "company", 
            to_object_id = child_id, 
            association_type = "13"
            )
        print("SUCCESS: ", parent_id, "is now parent to", child_id)
    except ApiException as e:
        print("Exception when calling associations_api->create: %s\n" % e)

def makeParent_companyToContact(company_id, contact_id):
    try:
        api_response = client.crm.companies.associations_api.create(
            company_id = company_id, 
            to_object_type = "contact", 
            to_object_id = contact_id, 
            association_type = "2"
            )
        print("SUCCESS: ", company_id, "is now associated with", contact_id)
    except ApiException as e:
        print("Exception when calling associations_api->create: %s\n" % e)

def makeParent_producttoContact(product_id, contact_id):
    try:
        api_response = client.crm.products.associations_api.create(
            product_id = product_id, 
            to_object_type = "contact", 
            to_object_id = contact_id, 
            association_type="associationType"
            )
        pprint(api_response)
    except ApiException as e:
        print("Exception when calling associations_api->create: %s\n" % e)

def in_dict(element, dictionary):
    print("element", element)
    if element in dictionary: # if company is not in dictionary then ...
        return True
    else:
        return False

def make_parents(associaton_type,  contact_directory, company_directory, associations_directory):
    print("make_parents has been called -",(time.strftime("%H:%M:%S", time.localtime())))
    associations_dict = data_to_dict("associations", associations_directory)
     
    if associaton_type == 'company-company':
        company_dict = data_to_dict("company", company_directory)
        failed = []
        open('failed.txt', 'w').close() # supposed to clear text in failed.txt before starting
        for line in associations_dict:
            try: # makes sure data is in right format
                parent = line
                child = associations_dict[line] 
                print("Parent:", parent, "Child:", child)
            except:
                print("ERROR: Check associations.xlsx")  

            if parent == "" or parent == ' ' or parent == None:
                print("Parent blank????", parent)
                continue
            elif child == "" or child == ' ' or child == None:
                print("Child blank????", child)
                continue
            else:
                if (isinstance(parent,list)) and isinstance(child,list): # if parent and children are both lists
                    parent_ids = [] # makes array for parent ids
                    child_ids = [] # makes array for child ids
                    for id in parent:
                        parent_ids.append(company_dict[id])

                    for id in parent: # if parent  is a list for every parent in the list. . .
                        if in_dict(id, company_dict): # check if parent exists in the dict
                            parent_ids.append(company_dict[id])
                        else: #if it does not exist
                            print("ERROR: Company not found in list of conatcts")
                            failed.append("Company:" + id + " was not found and failed to associate") # appends failure message to failed array
                    for id in child:
                        if in_dict(id, company_dict): # check if parent exists in the dict
                            child_ids.append(company_dict[id])
                        else: #if it does not exist
                            print("ERROR: Company not found in list of conatcts")
                            failed.append("Company:" + id + " was not found and failed to associate") # appends failure message to failed array
                    for parent_id in parent_id:
                        for child_id in child_ids:
                            makeParent_companyToCompany(parent_id, child_id) # makes association for every instance of child and parent    

                elif (isinstance(company_dict[parent],list)) and (not isinstance(child,list)): #if multiple instances of parent but not child
                    child_id = company_dict.get(child)
                    parent_ids = [] # makes array for parent ids
                    for parent in company_dict[parent]: # if company is a list for every company in the list. . .
                        if in_dict(parent, company_dict): # check if company exists in the dict
                            if isinstance(company_dict[parent], list):
                                parent_ids.extend(company_dict[parent])  
                            else:
                                parent_ids.append(company_dict[parent])  
                            
                        else: #if it does not exist
                            print("ERROR: Company not found in list of companies")
                            if isinstance(parent, list):
                                failed.append("Company:" + (', '.join(parent)) + " was not found and failed to associate") # appends failure message to failed array
                            else:
                                failed.append("Company: "+ parent + " was not found and failed to associate") # appends failure message to failed array

                    for parent_id in parent_ids: #makes association for every child in list
                        makeParent_companyToCompany(parent_id, child_id)   

                elif (isinstance(child,list)) and (not isinstance(company_dict[parent],list)): #if multiple instance of child but not multiple parents
                    parent_id = company_dict.get(parent)
                    child_ids = [] # makes array for child ids
                    for child in child_ids: # if child is a list for every company in the list. . .
                        if in_dict(child, company_dict): # check if child exists in the dict
                            if isinstance(company_dict[child], list):
                                child_ids.extend(company_dict[child])
                            else:
                                child_ids.append(company_dict[child])
                        else: #if it does not exist
                            print("ERROR: Company not found in list of conatcts")
                            if isinstance(child, list):
                                failed.append("Company:" + (', '.join(child)) + " was not found and failed to associate") # appends failure message to failed array
                            else:
                                failed.append("Company:" + child + " was not found and failed to associate") # appends failure message to failed array
                    
                    for child_id in child_ids: #makes association for every contact in list
                        makeParent_companyToCompany(parent_id, child_id)        

                elif company_dict.get(parent) == None:
                    print("ERROR: Company not found in list of companies")
                    failed.append("Company:" + parent + " was not found and failed to associate") # appends failure message to failed array
                elif company_dict.get(child) == None:
                    print("ERROR: Company not found in list of companies")
                    failed.append("Company:" + child + " was not found and failed to associate") # appends failure message to failed array
                else:
                    parent_id = company_dict.get(parent)
                    child_id = company_dict.get(child)
                    if isinstance(parent_id, list) and isinstance(child_id, list):
                        for parent_id in parent_id:
                            for child_id in child_id:
                                makeParent_companyToCompany(parent_id, child_id)

                    elif isinstance(parent_id, list) and (not isinstance(child_id, list)):
                        for parent_id in parent_id:
                            makeParent_companyToCompany(parent_id, child_id)

                    elif (not isinstance(parent_id, list)) and isinstance(child_id, list):
                        for child_id in child_id:
                            makeParent_companyToCompany(parent_id, child_id)
                    else:
                        makeParent_companyToCompany(parent_id, child_id)

        
    elif (associaton_type == 'company-contact') or (associaton_type == 'contact-company'):
        company_dict = data_to_dict("company", company_directory)
        contact_dict = data_to_dict("contact", contact_directory)
        failed = []
        open('failed.txt', 'w').close() # supposed to clear text in failed.txt before starting
        for line in associations_dict:
            try:
                company = line
                contact = associations_dict[line] # if there is a blank entery this will not work which is why its in try except
                print("Company:", company, "Contact:", contact)
            except:
                print("ERROR: Check associations.xlsx")

            if contact == "" or contact == ' ' or contact == None:
                print("contact blank????", contact)
                continue
            elif company == "" or company == ' ' or company == None:
                print("company blank????", company)
                continue
            else:
                if (isinstance(company_dict[company],list)) and isinstance(contact,list): # if both are lists . . .
                    company_ids = [] # makes array for company ids
                    contact_ids = [] # makes array for contact ids
                    for id in company_dict[company]: # if company is a list for every company in the list. . .
                        if in_dict(company, company_dict): # check if company exists in the dict
                            if isinstance(company_dict[company], list):
                                company_ids.extend(company_dict[company])  
                            else:
                                company_ids.append(company_dict[company])  
                            
                        else: #if it does not exist
                            print("ERROR: Company not found in list of companies", id)
                            if isinstance(company, list):
                                failed.append("Company:" + (', '.join(company)) + " was not found and failed to associate") # appends failure message to failed array
                            else:
                                failed.append("Company: "+ company + " was not found and failed to associate") # appends failure message to failed array

                    for id in contact: # if contact is a list for every company in the list. . .
                        if in_dict(id, contact_dict): # check if contact exists in the dict
                            if isinstance(contact_dict[id], list):
                                contact_ids.extend(contact_dict[id])
                            else:
                                contact_ids.append(contact_dict[id])
                        else: #if it does not exist
                            print("ERROR: Contact not found in list of conatcts")
                            if isinstance(contact, list):
                                failed.append("Contact:" + (', '.join(contact)) + " was not found and failed to associate") # appends failure message to failed array
                            else:
                                failed.append("Contact:" + contact + " was not found and failed to associate") # appends failure message to failed array
                    for company_id in company_ids:
                        for contact_id in contact_ids:
                            makeParent_companyToContact(company_id, contact_id) # makes association for every instance of child and parent

                elif (isinstance(company_dict[company],list)) and (not isinstance(contact,list)): #if multiple instances of company but not contact
                    contact_id = contact_dict.get(contact)
                    company_ids = [] # makes array for parent ids
                    for id in company_dict[company]: # if company is a list for every company in the list. . .
                        if in_dict(company, company_dict): # check if company exists in the dict
                            if isinstance(company_dict[company], list):
                                company_ids.extend(company_dict[company])  
                            else:
                                company_ids.append(company_dict[company])  
                            
                        else: #if it does not exist
                            print("ERROR: Company not found in list of companies")
                            if isinstance(company, list):
                                failed.append("Company:" + (', '.join(company)) + " was not found and failed to associate") # appends failure message to failed array
                            else:
                                failed.append("Company: "+ company + " was not found and failed to associate") # appends failure message to failed array

                    for company_id in company_ids: #makes association for every child in lsit
                        makeParent_companyToContact(company_id, contact_id)   

                elif (isinstance(contact,list)) and (not isinstance(company_dict[company],list)): #if multiple instance of contacts but not multiple companies
                    company_id = company_dict.get(company)
                    contact_ids = [] # makes array for contact ids
                    for id in contact: # if contact is a list for every company in the list. . .
                        if in_dict(id, contact_dict): # check if contact exists in the dict
                            if isinstance(contact_dict[id], list):
                                contact_ids.extend(contact_dict[id])
                            else:
                                contact_ids.append(contact_dict[id])
                        else: #if it does not exist
                            print("ERROR: Contact not found in list of conatcts")
                            if isinstance(contact, list):
                                failed.append("Contact:" + (', '.join(contact)) + " was not found and failed to associate") # appends failure message to failed array
                            else:
                                failed.append("Contact:" + contact + " was not found and failed to associate") # appends failure message to failed array
                    
                    for contact_id in contact_ids: #makes association for every contact in list
                        makeParent_companyToContact(company_id, contact_id)          
                elif company_dict.get(company) == None:
                    print("ERROR: Company not found in list of company")
                    failed.append("Company:" + company + " was not found and failed to associate") # appends failure message to failed array
                elif contact_dict.get(contact) == None:
                    print("ERROR: Contact not found in list of contacts")
                    failed.append("Contact:" + contact + " was not found and failed to associate") # appends failure message to failed array
                else:
                    company_id = company_dict.get(company)
                    contact_id = contact_dict.get(contact)
                    if isinstance(company_id, list) and isinstance(contact_id, list):
                        for company_id in company_id:
                            for contact_id in contact_id:
                                makeParent_companyToContact(company_id, contact_id)

                    elif isinstance(company_id, list) and (not isinstance(contact_id, list)):
                        for company_id in company_id:
                            makeParent_companyToContact(company_id, contact_id)

                    elif (not isinstance(company_id, list)) and isinstance(contact_id, list):
                        for contact_id in contact_id:
                            makeParent_companyToContact(company_id, contact_id)
                    else:
                        makeParent_companyToContact(company_id, contact_id)

        with open("failed.txt", 'w') as out:
            if failed: print("Some failures, check failed.txt for more info")
            for x in failed: # for every element in failed array
                out.write("".join(x) + "\n") # prints to failed.txt
    print("make_parents has been completed -",(time.strftime("%H:%M:%S", time.localtime())))

def handle_products(directory):
    wb = Workbook()
    
    wb = load_workbook(directory)
    ws = wb.active

    name_column = ws['A'] # names
    price_column = ws['B'] # prices
    description_column = ws['C'] 
    hs_sku__column = ws['D'] 
    hs_cost_of_goods_sold_column = ws['E'] 

    for (
        name, 
        price, 
        description, 
        hs_sku, 
        hs_cost_of_goods_sold
        ) in zip(
            name_column, 
            price_column, 
            description_column, 
            hs_sku__column,  
            hs_cost_of_goods_sold_column):
        # try:
            if name.value == "Name*": continue # doesnt count initial row
            if name.value == "Example" and price.value == "0": continue # doesnt count example
            if isinstance(price.value, str): (price.value).replace("$", "")

            properties = {}
            properties.update({"name": name.value})
            properties.update({"price": price.value})

            if description.value == None:
                description.value = ""
            else:
                properties["description"] = description.value

            if hs_sku.value == None:
                hs_sku.value = ""
            else:
                properties["hs_sku"] = hs_sku.value

            # if hs_recurring_billing_period.value == None:
            #     hs_recurring_billing_period.value = ""
            # else:
            #     properties["hs_recurring_billing_period"] = hs_recurring_billing_period.value

            if hs_cost_of_goods_sold.value == None:
                hs_cost_of_goods_sold.value = ""
            elif isinstance(hs_cost_of_goods_sold.value, str):
                if isinstance(hs_cost_of_goods_sold.value, str): (hs_cost_of_goods_sold.value).replace("$", "")
            else:
                properties["hs_cost_of_goods_sold"] = hs_cost_of_goods_sold.value

            #print(properties)
            import_product(properties)
        # except:
        #        print("Error with import") 
    

def import_product(properties):
    simple_public_object_input = SimplePublicObjectInput(properties=properties)
    try:
        api_response = client.crm.products.basic_api.create(simple_public_object_input=simple_public_object_input)
        pprint(api_response)
    except ApiException as e:
        print("Exception when calling basic_api->create: %s\n" % e)

def data_to_dict(type, directory): #takes data from xlsx file and returns it in a dictionary
    print('data_to_dict has been called -', (time.strftime("%H:%M:%S", time.localtime())))
    data = {}
    #create instance of workbook
    wb = Workbook()
    try:
        if type == 'company' or type == "associations":    
            #load exisiting work book
            wb = load_workbook(directory)
            
            # Create active worksheet
            ws = wb.active
            
            # Create variable for Columns
            id_column = ws['A'] # ids
            name_column = ws['B'] # names

            # iterates over 3 lists and executes 
            # 2 times as len(value)= 2 which is the
            # minimum among all the three 
            for (id, name) in zip(id_column, name_column):
                try:
                    if not (name.value.upper().isUpper()):
                        continue
                    elif name.value in data:
                        if isinstance(data[name.value], list):
                            arr = data[name.value]
                            arr.append(id.value)
                            data.update( {name.value : arr} )
                        else:
                            arr = [data[name.value]]
                            arr.append(id.value)
                            data.update( {name.value : arr} )
                    else:
                        data.update( {name.value : id.value} )
                except:
                        data.update( {name.value : id.value} ) 
                        
        elif type == 'contact':
            #load exisiting work book
            wb = load_workbook(directory)
            ws = wb.active
            
            id_column = ws['A'] #id
            first_name_column_b = ws['B'] # first name
            last_name_column_c = ws['C'] #last name

            for (id, first_name, last_name) in zip(id_column, first_name_column_b, last_name_column_c):
                if first_name.value == None and last_name.value == None:
                    full_name = ""
                elif first_name.value == None and last_name.value != None:
                    full_name = last_name.value
                elif first_name.value != None and last_name.value == None:
                    full_name = first_name.value
                else:
                    full_name = first_name.value + " " + last_name.value
                try:
                    if not (full_name.upper().isupper() ):
                        continue
                    elif full_name in data:
                        if isinstance(data[full_name], list):
                            arr = data[full_name]
                            arr.append(id.value)
                            data.update( {full_name : arr} )
                        else:
                            arr = [data[full_name]]
                            arr.append(id.value)
                            data.update( {full_name : arr} )
                    else:
                        data.update( {full_name : id.value} )
                except:
                        data.update( {full_name : id.value} )
        elif type == "products":
            #load exisiting work book
            wb = load_workbook(directory)
            
            # Create active worksheet
            ws = wb.active
            
            # Create variable for Columns
            id_column = ws['A'] # ids
            price_column = ws['C'] # ids
            name_column = ws['F'] # names

            # iterates over 3 lists and executes 
            # 2 times as len(value)= 2 which is the
            # minimum among all the three 
            for (id, name) in zip(id_column, name_column):
                try:
                    if not (full_name.upper().isupper() ):
                        continue
                    elif name.value in data:
                        if isinstance(data[name.value], list):
                            arr = data[name.value]
                            arr.append(id.value)
                            data.update( {name.value : arr} )
                        else:
                            arr = [data[name.value]]
                            arr.append(id.value)
                            data.update( {name.value : arr} )
                    else:
                        data.update( {name.value : id.value} )
                except:
                        data.update( {name.value : id.value} ) 

    except:
        print("ERROR: type error")
    
    print('data_to_dict has been completed -', (time.strftime("%H:%M:%S", time.localtime())))
    return data